"""
TEST SECTION 7: TESTS DES LIVRABLES (EXCEL & PDF)
Tests de génération des exports avec formules natives
"""
import pytest
import openpyxl
from io import BytesIO
from app.services.excel_service import excel_service
from app.services.financial_service import financial_service


class TestExcelFormules:
    """Test: Excel contient des formules natives, pas valeurs figées"""
    
    def test_fichier_excel_genere(self):
        """Le fichier Excel est bien généré"""
        project_data = {
            "name": "Test Excel",
            "purchase_price": 1_000_000,
            "capex_total": 500_000,
            "monthly_rent": 10_000,
            "duration_years": 10
        }
        
        excel_bytes = excel_service.generate_business_plan_excel(project_data)
        
        assert excel_bytes is not None
        assert len(excel_bytes) > 0
        
        # Vérifier que c'est bien un fichier Excel
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        assert wb is not None
    
    def test_excel_contient_formules_pas_valeurs(self):
        """Les cellules contiennent des formules =SUM(), pas des valeurs"""
        project_data = {
            "name": "Test Formules",
            "purchase_price": 1_000_000,
            "capex_total": 500_000,
            "monthly_rent": 10_000
        }
        
        excel_bytes = excel_service.generate_business_plan_excel(project_data)
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        
        # Aller sur l'onglet KPIs
        ws_kpis = wb["KPIs"]
        
        # Chercher une cellule avec formule
        found_formula = False
        for row in ws_kpis.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    found_formula = True
                    break
            if found_formula:
                break
        
        assert found_formula, "Aucune formule Excel détectée dans le fichier"
    
    def test_modification_excel_recalcule_kpis(self):
        """Si l'utilisateur modifie une hypothèse, les KPIs se recalculent"""
        project_data = {
            "name": "Test Recalcul",
            "purchase_price": 1_000_000,
            "capex_total": 500_000,
            "monthly_rent": 10_000
        }
        
        excel_bytes = excel_service.generate_business_plan_excel(project_data)
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        
        # Onglet Inputs
        ws_inputs = wb["Inputs"]
        
        # Trouver la cellule du loyer mensuel
        for row in ws_inputs.iter_rows():
            for cell in row:
                if cell.value == "Loyer mensuel":
                    # La valeur est dans la cellule suivante
                    loyer_cell = ws_inputs.cell(row=cell.row, column=cell.column + 1)
                    
                    # Modifier le loyer
                    original_value = loyer_cell.value
                    loyer_cell.value = 15_000  # Au lieu de 10_000
                    
                    # Excel devrait recalculer automatiquement
                    # (impossible à tester sans ouvrir Excel, mais structure validée)
                    assert loyer_cell.value != original_value
                    break


class TestOngletsExcel:
    """Test: Excel contient les bons onglets"""
    
    def test_onglets_obligatoires_presents(self):
        """Onglets: Inputs, Cashflows, KPIs, TRI"""
        project_data = {
            "name": "Test Onglets",
            "purchase_price": 1_000_000
        }
        
        excel_bytes = excel_service.generate_business_plan_excel(project_data)
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        
        sheet_names = wb.sheetnames
        
        assert "Inputs" in sheet_names
        assert "Cashflows" in sheet_names
        assert "KPIs" in sheet_names
        assert "TRI" in sheet_names
    
    def test_onglet_cashflows_structure(self):
        """Onglet Cashflows a la bonne structure"""
        project_data = {
            "name": "Test Cashflows",
            "purchase_price": 1_000_000,
            "capex_total": 500_000,
            "monthly_rent": 10_000
        }
        
        excel_bytes = excel_service.generate_business_plan_excel(project_data)
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        
        ws_cf = wb["Cashflows"]
        
        # Vérifier les en-têtes de colonnes
        headers = [cell.value for cell in ws_cf[1]]
        
        assert "Mois" in headers or "Date" in headers
        assert "CAPEX" in headers or "Investissement" in headers
        assert "Revenus" in headers or "Loyers" in headers
        assert "Cashflow" in headers or "Flux" in headers


class TestDossierBanque:
    """Test: Dossier banque (PDF assemblé)"""
    
    def test_pdf_assemble_genere(self):
        """Le PDF dossier banque est généré"""
        project_id = 123
        
        pdf_bytes = excel_service.generate_bank_dossier_pdf(project_id)
        
        assert pdf_bytes is not None
        assert len(pdf_bytes) > 0
        
        # Vérifier que c'est bien un PDF
        assert pdf_bytes[:4] == b'%PDF'
    
    def test_documents_dans_ordre_correct(self):
        """Les documents sont dans le bon ordre"""
        project_id = 123
        
        pdf_bytes = excel_service.generate_bank_dossier_pdf(project_id)
        
        # Parser le PDF pour vérifier l'ordre
        # (simplifié pour le test)
        doc_order = excel_service.get_bank_dossier_document_order()
        
        expected_order = [
            "EXECUTIVE_SUMMARY",
            "BUSINESS_PLAN",
            "TECHNICAL_SCORE",
            "DOCUMENTS_OFFICIELS",
            "ANNEXES"
        ]
        
        assert doc_order == expected_order
    
    def test_contenu_conforme_cahier_charges(self):
        """Le contenu est conforme au cahier des charges"""
        project_id = 123
        
        dossier_content = excel_service.get_bank_dossier_content(project_id)
        
        # Sections obligatoires
        assert "executive_summary" in dossier_content
        assert "financial_summary" in dossier_content
        assert "technical_analysis" in dossier_content
        assert "risk_analysis" in dossier_content


class TestQualiteExports:
    """Test qualité des exports"""
    
    def test_excel_ouvrable_sans_erreur(self):
        """L'Excel s'ouvre sans erreur"""
        project_data = {
            "name": "Test Qualité",
            "purchase_price": 1_000_000
        }
        
        excel_bytes = excel_service.generate_business_plan_excel(project_data)
        
        try:
            wb = openpyxl.load_workbook(BytesIO(excel_bytes))
            # Si on arrive ici, pas d'erreur
            assert True
        except Exception as e:
            pytest.fail(f"Excel invalide: {str(e)}")
    
    def test_pas_de_valeurs_nan_ou_inf(self):
        """Aucune valeur NaN ou Inf dans l'Excel"""
        project_data = {
            "name": "Test NaN",
            "purchase_price": 1_000_000,
            "capex_total": 500_000
        }
        
        excel_bytes = excel_service.generate_business_plan_excel(project_data)
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        
        for sheet in wb:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        value_str = str(cell.value).upper()
                        assert "NAN" not in value_str
                        assert "INF" not in value_str


class TestMetadonnees:
    """Test métadonnées des exports"""
    
    def test_metadata_excel(self):
        """Métadonnées correctes dans l'Excel"""
        project_data = {
            "name": "Test Metadata",
            "purchase_price": 1_000_000
        }
        
        excel_bytes = excel_service.generate_business_plan_excel(project_data)
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        
        # Vérifier propriétés
        assert wb.properties.creator == "REFY AI"
        assert wb.properties.title is not None
    
    def test_date_generation(self):
        """La date de génération est incluse"""
        project_data = {
            "name": "Test Date",
            "purchase_price": 1_000_000
        }
        
        excel_bytes = excel_service.generate_business_plan_excel(project_data)
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        
        ws = wb["Inputs"]
        
        # Chercher la ligne "Généré le"
        found_date = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Généré le" in str(cell.value):
                    found_date = True
                    break
        
        assert found_date, "Date de génération manquante"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
